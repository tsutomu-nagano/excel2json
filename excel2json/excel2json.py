import yaml
import openpyxl
import pandas as pd
import re
import json
import unicodedata
import sys
import copy

import exceptions

class Excel2JSON():


    @staticmethod
    def __getcell(name, wb, info):

        __num_ptn = re.compile("^[0-9]+")

        sheet = info["sheet"]
        address = info["address"]
        value = str(wb[sheet][address].value).strip()
        value = "" if value == "None" else value

        if "require" in info:
            if value == "":
                raise exceptions.MissingRequiredError(name)

        if "numeric" in info:
            if __num_ptn.match(value) is None:
                raise exceptions.InvalidNumericValueError(name)


        return(str(wb[sheet][address].value))



    @staticmethod
    def __rename_header(df, colmaps):

        # 改行コードは削除
        df.columns = [re.sub(r"(\r\n|\n|\r)", "", column) for column in df.columns.values]

        # 文字列一致で置換
        df = df.rename(columns = {colmap["ja"]["text"]:colmap["var"] for colmap in colmaps if "text" in colmap["ja"]})

        # 正規表現で置換
        headers  = df.columns.values
        patterns = {colmap["ja"]["re"]:colmap["var"] for colmap in colmaps if "re" in colmap["ja"]}
        if len(patterns) >= 1:
            for pattern, var in patterns.items():
                
                headers = [re.sub(pattern, var, h) for h in headers]
            
            df.columns = headers

        return(df)



    @staticmethod
    def __get_missing_require_fields(df, colmaps):

        names_exists = {colmap["var"]:False for colmap in colmaps if "require" in colmap}

        [names_exists.update((column, True)) for column in df.columns.values if column in names_exists]

        return([name for name, exist in names_exists.items() if not exist])


    @staticmethod
    def __gettable(src, info):


        sheet = info["sheet"]
        address = info["address"]

        m = re.match("(?P<col>[A-Z]+)(?P<row>[0-9]+)", address)
        srow = int(m.group("row"))


        # 使用するヘッダーを決める処理
        df = pd.read_excel(src, sheet_name=sheet, skiprows=srow - 2, header = None)
        usecols = [index for index, value in enumerate(df.loc[1].tolist()) if not pd.isna(value)]
        
        df = pd.read_excel(src, sheet_name=sheet, skiprows=srow - 1,  usecols = usecols, dtype=str
                        ).fillna("")
        

        # ヘッダーの処理
        if "colmap" in info:
            colmaps = info["colmap"]

            # 名前変換
            df = Excel2JSON.__rename_header(df, colmaps)

            # 必須項目のチェック
            missing_require_fields = Excel2JSON.__get_missing_require_fields(df, colmaps)
            if len(missing_require_fields) >= 1:
                raise exceptions.MissingRequiredError(",".join(missing_require_fields))

            # 値のフォーマット
            formats = {colmap["var"]:colmap["format"] for colmap in colmaps if "format" in colmap and colmap["var"] in df.columns}

            for col_var, format in formats.items():
                if "wide_to_narrow" in format:
                    df[col_var] = df[col_var].map(lambda x : unicodedata.normalize("NFKC", x))



        return([row.to_dict() for index, row in df.iterrows()])


    @staticmethod
    def __getlist(wb, src, info):

        sheet_ptn = info["sheet"]

        # for sheet in wb.sheetnames:
        #     if re.match(sheet_ptn, sheet):
        #         print(sheet)
        #         print(copy.deepcopy(info["listitem"]))
        #         print(Excel2JSON.__xls_with_yaml2json_core(copy.deepcopy(info["listitem"]), wb, src, sheet))

        items = [Excel2JSON.__xls_with_yaml2json_core(copy.deepcopy(info["listitem"]), wb, src, sheet) for sheet in wb.sheetnames if re.match(sheet_ptn, sheet)]

        return(items)


    @staticmethod
    def __xls_with_yaml2json_core(y, wb, src, sheet = "", name = ""):
        if isinstance(y, dict):
            if "type" in y:
                conv_type = y["type"]

                if sheet != "":
                    y["sheet"] = sheet

                if conv_type == "cell":
                    y = Excel2JSON.__getcell(name, wb, y)
                
                if conv_type == "table":
                    y = Excel2JSON.__gettable(src, y)

                if conv_type == "list":
                    y = Excel2JSON.__getlist(wb, src, y)

            else:
                for k in y:
                    y[k] = Excel2JSON.__xls_with_yaml2json_core(y[k], wb, src, sheet, name = k)

        return(y)




    @classmethod
    def xls_with_yaml2json(cls, src, config):
        with open(config, 'r') as ymlf:
            json_ = cls.__xls_with_yaml2json_core(
                y = yaml.safe_load(ymlf),
                wb = openpyxl.load_workbook(src),
                src = src
            )
            return(json_)


if __name__ == "__main__":
    args = sys.argv

    src = args[1]
    config = args[2]
    dest = args[3]

    data = Excel2JSON.xls_with_yaml2json(src, config)
    
    with open(dest, 'w') as f:
        json.dump(data, f, indent=2)

